"""
Экспорт результатов скоринга в xlsx.

Создаёт файл с:
  - Отсортированной по баллу таблицей
  - Настоящими гиперссылками в колонках «Закупка» и «Площадка»
  - Условным форматированием по баллу (зелёный → красный)
  - Замороженным первым рядом
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


# Колонки результата: (заголовок, поле в словаре, ширина)
EXPORT_COLUMNS = [
    ("Балл",         "score",            8),
    ("Закупка",      "number",           18),
    ("Название",     "name",             55),
    ("НМЦ",          "price",            14),
    ("Площадка",     "platform",         18),
    ("Тип торгов",   "trade_type",       18),
    ("Способ отбора","selection_method", 26),
    ("Регион",       "region",           20),
    ("Дедлайн",      "deadline",         18),
    ("Дней до дедл.","days_to_deadline", 14),
]


def _score_color(score: float | None) -> str:
    """HEX-цвет заливки ячейки балла."""
    if score is None:
        return "DDDDDD"
    if score >= 0.7:
        return "6FCF97"   # зелёный
    if score >= 0.4:
        return "F2C94C"   # жёлтый
    return "EB5757"        # красный


def _fmt_price(price: float | None) -> str:
    if price is None:
        return "—"
    return f"{price:,.0f} ₽".replace(",", "\u202f")


def _fmt_date(dt: datetime | None) -> str:
    if dt is None:
        return "—"
    return dt.strftime("%d.%m.%Y %H:%M")


def export_to_xlsx(records: list[dict]) -> bytes:
    """
    Принимает список записей (уже отсортированных scorer'ом),
    возвращает байты xlsx-файла.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Результаты"

    # --- Стили ---
    header_fill = PatternFill("solid", fgColor="1A1A2E")
    header_font = Font(bold=True, color="EAEAEA", size=10)
    thin = Side(style="thin", color="CCCCCC")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    link_font = Font(color="2F80ED", underline="single", size=10)
    data_font = Font(size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # --- Заголовок ---
    for col_idx, (header, _, width) in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = cell_border
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # --- Данные ---
    for row_idx, record in enumerate(records, start=2):
        score = record.get("score")

        for col_idx, (_, field, _) in enumerate(EXPORT_COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = cell_border
            cell.font = data_font

            if field == "score":
                cell.value = score
                cell.number_format = "0.00"
                cell.alignment = center
                cell.fill = PatternFill("solid", fgColor=_score_color(score))
                cell.font = Font(bold=True, size=10)

            elif field == "number":
                cell.value = record.get("number") or "—"
                url = record.get("card_url")
                if url:
                    cell.hyperlink = url
                    cell.font = link_font
                else:
                    cell.font = data_font
                cell.alignment = center

            elif field == "platform":
                cell.value = record.get("platform") or "—"
                url = record.get("platform_url")
                if url:
                    cell.hyperlink = url
                    cell.font = link_font
                else:
                    cell.font = data_font
                cell.alignment = center

            elif field == "price":
                cell.value = _fmt_price(record.get("price"))
                cell.alignment = center

            elif field == "deadline":
                cell.value = _fmt_date(record.get("deadline"))
                cell.alignment = center

            elif field == "days_to_deadline":
                v = record.get("days_to_deadline")
                cell.value = v if v is not None else "—"
                cell.alignment = center
                if isinstance(v, int) and v < 3:
                    cell.font = Font(bold=True, color="EB5757", size=10)

            elif field == "name":
                cell.value = record.get("name") or "—"
                cell.alignment = left

            else:
                cell.value = record.get(field) or "—"
                cell.alignment = left

        ws.row_dimensions[row_idx].height = 18

    # --- Автофильтр ---
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
