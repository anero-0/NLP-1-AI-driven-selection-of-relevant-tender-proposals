"""
Парсер xlsx-выгрузок из Контур.Закупки.

Особенности формата:
  - Строка 1: группы колонок (Закупка, НМЦ, Заказчик …) — пропускаем
  - Строка 2: имена колонок — используем как заголовки
  - Строки 3+: данные
  - Ячейки «Номер» и «ЭТП» содержат гиперссылки (cell.hyperlink.target),
    которые pandas теряет — поэтому читаем через openpyxl
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any

import openpyxl

# Обязательные колонки — если хотя бы одной нет, файл считается невалидным
REQUIRED_COLUMNS = {"Номер", "Название", "ЭТП"}

# Маппинг: имя колонки в файле → внутреннее имя поля
COLUMN_MAP = {
    "Номер": "number",
    "Название": "name",
    "НМЦ": "price",
    "Аванс": "advance",
    "Валюта закупки": "currency",
    "Дата публикации": "pub_date",
    "Планируемая дата публикации": "planned_pub_date",
    "Окончание приема заявок": "deadline",
    "Проведение отбора": "selection_date",
    "Этап отбора": "selection_stage",
    "Тип торгов": "trade_type",
    "Ссылка на ЕИС": "eis_link",
    "Способ отбора": "selection_method",
    "ЭТП": "platform",
    "СМП, СОНО": "smp",
    "Размещает  закупку": "placer",
    "Контактное лицо": "contact",
    "Метка ": "tag",
    "Комментарий": "comment",
    "Ответственный": "responsible",
    "Регион": "region",
    # Дублирующийся «Название» заказчика (col 22 в файле)
    # openpyxl даст нам оба — обработаем по позиции
}


class ParseError(ValueError):
    """Понятная ошибка парсинга файла."""
    pass


def _safe_float(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return float(str(value).replace("\xa0", "").replace(" ", "").replace(",", "."))
    except (ValueError, TypeError):
        return None


def _safe_datetime(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    try:
        return datetime.fromisoformat(str(value))
    except (ValueError, TypeError):
        return None


def parse_xlsx(file_bytes: bytes) -> list[dict]:
    """
    Парсит xlsx-файл Контур.Закупки и возвращает список записей.

    Каждая запись — словарь с полями:
        number, name, price, pub_date, deadline, trade_type,
        selection_method, platform, region,
        card_url  (гиперссылка ячейки «Номер»),
        platform_url  (гиперссылка ячейки «ЭТП»),
        days_to_deadline  (вычисляемое)

    Raises:
        ParseError — если файл нечитаемый, пустой или отсутствуют нужные колонки.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ParseError(
            f"Не удалось прочитать файл. Убедитесь, что это корректный .xlsx. "
            f"Детали: {exc}"
        ) from exc

    ws = wb.active

    if ws.max_row < 3:
        raise ParseError(
            "Файл пустой или не содержит строк с данными (ожидается минимум 3 строки)."
        )

    # --- Читаем строку 2 как заголовки ---
    headers: list[str | None] = [
        ws.cell(row=2, column=col).value for col in range(1, ws.max_column + 1)
    ]

    # Нормализуем: убираем пробелы, None оставляем
    headers = [h.strip() if isinstance(h, str) else h for h in headers]

    # Проверяем обязательные колонки
    present = {h for h in headers if h}
    missing = REQUIRED_COLUMNS - present
    if missing:
        raise ParseError(
            f"Не найдены обязательные колонки: {', '.join(sorted(missing))}. "
            f"Убедитесь, что загружаете выгрузку из Контур.Закупки."
        )

    # Индексы нужных колонок (0-based внутри headers)
    col_index: dict[str, int] = {}
    for idx, h in enumerate(headers):
        if h and h not in col_index:  # берём первое вхождение (для дублей)
            col_index[h] = idx
    # Для второго «Название» (заказчик): ищем второе вхождение
    customer_name_col = None
    name_count = 0
    for idx, h in enumerate(headers):
        if h == "Название":
            name_count += 1
            if name_count == 2:
                customer_name_col = idx
                break

    # --- Читаем данные, строки 3+ ---
    records: list[dict] = []
    now = datetime.now()

    for row_num in range(3, ws.max_row + 1):
        row_cells = [ws.cell(row=row_num, column=col) for col in range(1, ws.max_column + 1)]

        def cell_val(col_name: str) -> Any:
            idx = col_index.get(col_name)
            return row_cells[idx].value if idx is not None else None

        def cell_link(col_name: str) -> str | None:
            idx = col_index.get(col_name)
            if idx is None:
                return None
            cell = row_cells[idx]
            if cell.hyperlink:
                return cell.hyperlink.target
            return None

        number = cell_val("Номер")
        if not number:
            continue  # пропускаем пустые строки

        name = cell_val("Название") or ""
        pub_date = _safe_datetime(cell_val("Дата публикации"))
        deadline = _safe_datetime(cell_val("Окончание приема заявок"))

        # Дни от публикации до дедлайна
        days_to_deadline: int | None = None
        if pub_date and deadline:
            days_to_deadline = (deadline - pub_date).days
        elif deadline:
            days_to_deadline = (deadline - now).days

        # Имя заказчика из второго столбца «Название»
        customer_name = None
        if customer_name_col is not None:
            customer_name = row_cells[customer_name_col].value

        record = {
            "number": str(number).strip(),
            "card_url": cell_link("Номер"),
            "name": str(name).strip(),
            "price": _safe_float(cell_val("НМЦ")),
            "has_price": _safe_float(cell_val("НМЦ")) is not None,
            "advance": _safe_float(cell_val("Аванс")),
            "currency": cell_val("Валюта закупки"),
            "pub_date": pub_date,
            "deadline": deadline,
            "days_to_deadline": days_to_deadline,
            "trade_type": cell_val("Тип торгов"),
            "selection_method": cell_val("Способ отбора"),
            "selection_stage": cell_val("Этап отбора"),
            "platform": str(cell_val("ЭТП") or "").strip() or None,
            "platform_url": cell_link("ЭТП"),
            "region": cell_val("Регион"),
            "customer_name": str(customer_name).strip() if customer_name else None,
            "contact": cell_val("Контактное лицо"),
            "tag": cell_val("Метка "),
            "comment": cell_val("Комментарий"),
            "score": None,  # будет заполнен скорером
        }
        records.append(record)

    if not records:
        raise ParseError("Файл не содержит строк с данными (все строки пустые).")

    return records
